#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
配置管理器 —— 加载并验证 JSON 格式的股票配置

配置文件 (config.json) 各字段说明：

最简写法（一行一个股票名称，参数全局共享）：
{
    "stocks": ["神火股份", "贵州茅台"],       // 直接写名称数组
    "start_date": "2020-01-01",               // 全局起始日期
    "end_date": "2026-05-19",                 // 全局截止日期
    "analysis": {                              // 全局分析参数（可选）
        "grid_levels": 4,
        "atr_multiplier": 0.5,
        "t_single_qty": 500
    }
}

标准写法（每只股票独立设置，代码/文件名可手动指定）：
{
    "stocks": [
        {
            "stock_name": "神火股份",
            "stock_code": "000933",           // 不填则自动查询
            "start_date": "2022-04-19",
            "end_date": "2026-05-18",
            "output_file": "神火股份.xlsx",    // 不填则自动生成
            "analysis": { ... }
        }
    ]
}
{
    "stocks": [
        {
            "stock_name": "神火股份",         // 股票名称
            "stock_code": "000933",           // 股票代码（不填则自动查询）
            "start_date": "2022-04-19",       // 数据起始日期
            "end_date":   "2026-05-18",       // 数据截止日期
            "output_file": "神火股份.xlsx",    // 输出文件名（不填则自动生成）
            "analysis": {                     // 分析参数（可选）
                "grid_levels": 4,             // 网格交易档位数
                "atr_multiplier": 0.5,        // 网格间距 = ATR × 此系数
                "t_single_qty": 500           // 每笔做T股数，不设则自动估算
            }
        }
    ]
}
"""

import json
import os
from typing import Dict, List, Optional
import httpx


class ConfigError(Exception):
    """配置加载或验证错误"""
    pass

# ── 东方财富搜索API ──
SEARCH_API_URL = "https://searchadapter.eastmoney.com/api/suggest/get"
SEARCH_TOKEN = "D43BF722C8E33C0E8E32BADB5C16D1D4"


def lookup_stock_code(stock_name: str) -> Optional[str]:
    """
    通过股票名称查询股票代码

    Args:
        stock_name: 股票名称，如 "神火股份"

    Returns:
        带交易所后缀的股票代码，如 "000933.SZ"；查不到返回 None
    """
    try:
        with httpx.Client(timeout=10.0) as client:
            resp = client.get(SEARCH_API_URL, params={
                "input": stock_name,
                "type": 14,
                "token": SEARCH_TOKEN,
            })
            resp.raise_for_status()
            data = resp.json()
    except Exception as e:
        print(f"  [警告] 股票代码查询失败({stock_name}): {e}")
        return None

    items = data.get("QuotationCodeTable", {}).get("Data", [])
    if not items:
        return None

    # 精确匹配名称（取第一个）
    code = items[0].get("Code", "")
    if not code:
        return None

    # 判断交易所
    sec_type = items[0].get("SecurityTypeName", "")
    if "沪" in sec_type or code.startswith(("6", "9")):
        return f"{code}.SH"
    else:
        return f"{code}.SZ"


def load_config(config_path: str) -> List[Dict]:
    """
    加载并验证 JSON 配置文件（仅股票配置列表）

    Returns:
        标准化后的股票配置列表
    """
    stocks, _ = load_config_full(config_path)
    return stocks


def load_config_full(config_path: str):
    """
    加载完整配置，返回 (股票配置列表, 全局配置字典)

    全局配置字段（可选）：
      - t_start_date / t_end_date: 波段T分析起止日期 (如 "2026-01-01" / "2026-05-21")
      - t_analysis_date: （已废弃，仅向下兼容）做T分析日期或范围
    """
    if not os.path.exists(config_path):
        raise ConfigError(f"配置文件不存在: {config_path}")

    # ── 检测文件类型：.xlsx 走 Excel 读取，否则走 JSON ──
    ext = os.path.splitext(config_path)[1].lower()
    if ext == ".xlsx":
        return _load_config_from_excel(config_path)

    try:
        with open(config_path, "r", encoding="utf-8") as f:
            raw = json.load(f)
    except json.JSONDecodeError as e:
        raise ConfigError(f"JSON 解析失败: {e}")

    # 根节点必须包含 stocks 数组
    if not isinstance(raw, dict) or "stocks" not in raw:
        raise ConfigError("配置文件根节点必须包含 'stocks' 数组")

    stocks_raw = raw["stocks"]
    if not isinstance(stocks_raw, list) or len(stocks_raw) == 0:
        raise ConfigError("'stocks' 必须是非空数组")

    # 判断 stocks 格式：简单字符串数组 vs 对象数组
    is_simple_list = isinstance(stocks_raw[0], str)

    # 从根节点读取全局默认值（简单数组格式时使用）
    default_start = raw.get("start_date", "")
    default_end = raw.get("end_date", "")
    default_analysis = raw.get("analysis")

    validated = []

    for idx, item in enumerate(stocks_raw, 1):
        if is_simple_list:
            # 简单数组：元素就是股票名称
            stock_name = item
            stock_code = ""
            output_file = ""
            start_date = default_start
            end_date = default_end
            analysis = default_analysis
        else:
            # 对象数组：原有逻辑
            stock_name = item.get("stock_name", "")
            stock_code = item.get("stock_code", "")
            output_file = item.get("output_file", "")
            start_date = item.get("start_date", "")
            end_date = item.get("end_date", "")
            analysis = item.get("analysis")

        # ── 若没有股票代码，尝试通过名称自动查询 ──
        if not stock_code:
            if not stock_name:
                raise ConfigError(f"第 {idx} 只股票: stock_name 和 stock_code 至少填一个")
            print(f"  [自动查询] 正在查询「{stock_name}」的股票代码...")
            stock_code = lookup_stock_code(stock_name)
            if not stock_code:
                raise ConfigError(f"第 {idx} 只股票: 未能自动查询到「{stock_name}」的代码，请手动填写 stock_code")
            print(f"    -> 查到代码: {stock_code}")
        else:
            # 已有代码，标准化
            stock_code = str(stock_code).strip()
            if not stock_code.endswith(".SZ") and not stock_code.endswith(".SH"):
                if stock_code.startswith(("6", "9")):
                    stock_code = f"{stock_code}.SH"
                else:
                    stock_code = f"{stock_code}.SZ"

        # ── 若没有输出文件名，自动生成 ──
        if not output_file:
            name_part = stock_name if stock_name else stock_code.replace(".SZ", "").replace(".SH", "")
            output_file = f"{name_part}_{stock_code}抓取数据.xlsx"

        validated.append({
            "stock_name": stock_name,
            "stock_code": stock_code,
            "start_date": start_date,
            "end_date": end_date,
            "output_file": output_file,
            "analysis": analysis,
        })

    # 提取全局配置
    global_config = {}
    # 支持 t_start_date / t_end_date 单独配置（替代旧的 t_analysis_date 范围写法）
    ts = raw.get("t_start_date", "")
    te = raw.get("t_end_date", "")
    if ts and te:
        from datetime import datetime, timedelta
        try:
            sd = datetime.strptime(ts.strip(), "%Y-%m-%d")
            ed = datetime.strptime(te.strip(), "%Y-%m-%d")
            if sd > ed:
                sd, ed = ed, sd
            dates = []
            cur = sd
            while cur <= ed:
                dates.append(cur.strftime("%Y-%m-%d"))
                cur += timedelta(days=1)
            global_config["t_start_date"] = ts.strip()
            global_config["t_end_date"] = te.strip()
            global_config["_t_dates"] = dates
        except ValueError:
            pass
    # 向下兼容：仍支持旧的 t_analysis_date 字段
    if "_t_dates" not in global_config and "t_analysis_date" in raw:
        global_config["_t_dates"] = parse_analysis_dates(raw["t_analysis_date"])

    return validated, global_config


def _load_config_from_excel(excel_path: str):
    """
    从 my_stocks.xlsx 加载配置

    Excel 列说明：
      stocks       - 股票名称（必填）
      start_date   - 数据起始日期
      end_date     - 数据截止日期
      t_start_date - 波段T分析起始日期
      t_end_date   - 波段T分析截止日期
      enable       - 是否分析该股票（1=分析，0=跳过，默认分析）

    每行一只股票，各股票可独立设置日期范围。
    """
    import openpyxl
    from datetime import datetime

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # 读取表头
    headers = [cell.value for cell in ws[1]]
    if not headers or headers[0] != "stocks":
        raise ConfigError("Excel 第一列必须是 'stocks'（股票名称）")

    # 构建列名→列号映射
    col_map = {}
    for i, h in enumerate(headers):
        if h:
            col_map[str(h).strip()] = i

    def _cell_str(row, col_idx):
        """取单元格内容并转为字符串"""
        val = row[col_idx] if col_idx < len(row) else None
        if val is None:
            return ""
        if isinstance(val, datetime):
            return val.strftime("%Y-%m-%d")
        return str(val).strip()

    validated = []
    global_config = {}
    t_start_vals = []
    t_end_vals = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        stock_name = _cell_str(row, col_map.get("stocks", 0))
        if not stock_name:
            continue  # 跳过空行

        # ── 检查是否启用 ──
        enable_col = col_map.get("enable", 5)
        enable_val = row[enable_col] if enable_col < len(row) else None
        if enable_val is not None and str(enable_val).strip() == "0":
            print(f"  [跳过] {stock_name} 未启用（enable=0），跳过")
            continue

        start_date = _cell_str(row, col_map.get("start_date", 1))
        end_date = _cell_str(row, col_map.get("end_date", 2))

        # ── 自动查询股票代码 ──
        stock_code = ""
        print(f"  [自动查询] 正在查询「{stock_name}」的股票代码...")
        stock_code = lookup_stock_code(stock_name)
        if not stock_code:
            print(f"  [警告] 未能自动查询到「{stock_name}」的代码，跳过该股票")
            continue
        print(f"    -> 查到代码: {stock_code}")

        # 收集 T 分析日期（取所有行的最小/最大）
        ts = _cell_str(row, col_map.get("t_start_date", 3))
        te = _cell_str(row, col_map.get("t_end_date", 4))
        if ts:
            t_start_vals.append(ts)
        if te:
            t_end_vals.append(te)

        # ── 自动生成输出文件名 ──
        name_part = stock_name if stock_name else stock_code.replace(".SZ", "").replace(".SH", "")
        output_file = f"{name_part}_{stock_code}抓取数据.xlsx"

        validated.append({
            "stock_name": stock_name,
            "stock_code": stock_code,
            "start_date": start_date,
            "end_date": end_date,
            "output_file": output_file,
            "analysis": None,
        })

    if not validated:
        raise ConfigError("Excel 中未找到有效的股票数据")

    # 从所有行中取 t_start_date 的最小值和 t_end_date 的最大值
    if t_start_vals and t_end_vals:
        t_start = min(t_start_vals)
        t_end = max(t_end_vals)
        global_config["t_start_date"] = t_start
        global_config["t_end_date"] = t_end
        # 生成日期列表
        from datetime import datetime, timedelta
        try:
            sd = datetime.strptime(t_start, "%Y-%m-%d")
            ed = datetime.strptime(t_end, "%Y-%m-%d")
            dates = []
            cur = sd
            while cur <= ed:
                dates.append(cur.strftime("%Y-%m-%d"))
                cur += timedelta(days=1)
            global_config["_t_dates"] = dates
        except ValueError:
            pass

    return validated, global_config


def parse_analysis_dates(raw_value) -> List[str]:
    """
    解析 t_analysis_date 配置，返回日期字符串列表

    支持格式：
      - 单个日期: "2026-05-19"
      - 日期范围: "2026-01-01~2026-01-31" 或 "2026-01-01 ~ 2026-01-31"
      - 日期列表: ["2026-01-01", "2026-01-15"]
    """
    if raw_value is None:
        return []

    # 列表格式
    if isinstance(raw_value, list):
        return [str(d).strip() for d in raw_value if str(d).strip()]

    # 字符串格式
    s = str(raw_value).strip()
    if not s:
        return []

    # 范围格式: "2026-01-01~2026-01-31"
    from datetime import datetime, timedelta
    for sep in ["~", "～"]:
        if sep in s:
            parts = s.split(sep)
            if len(parts) == 2:
                a, b = parts[0].strip(), parts[1].strip()
                if a and b:
                    try:
                        sd = datetime.strptime(a, "%Y-%m-%d")
                        ed = datetime.strptime(b, "%Y-%m-%d")
                        if sd > ed:
                            sd, ed = ed, sd
                        dates = []
                        cur = sd
                        while cur <= ed:
                            dates.append(cur.strftime("%Y-%m-%d"))
                            cur += timedelta(days=1)
                        return dates
                    except ValueError:
                        pass
            break

    # 单个日期
    return [s]


def print_config_list(configs: List[Dict]) -> None:
    """
    打印配置列表摘要到控制台

    Args:
        configs: load_config() 返回的配置列表
    """
    print(f"\n{'=' * 50}")
    print(f"[配置] 共 {len(configs)} 只股票待抓取")
    print(f"{'=' * 50}")

    for idx, cfg in enumerate(configs, 1):
        name = cfg.get("stock_name", "未知")
        code = cfg["stock_code"]
        start = cfg.get("start_date") or "不限"
        end = cfg.get("end_date") or "不限"
        out = cfg["output_file"]
        print(f"  [{idx}] {name} ({code})")
        print(f"      时间: {start} ~ {end}")
        print(f"      输出: {out}")

    print(f"{'=' * 50}")
